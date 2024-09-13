# -*- coding: utf-8 -*-
"""
Created on Fri Sep  6 16:50:03 2024

@author: ED
"""

# -----------------------------------------------------------------------------------------------#
                                    # IMPORTAÇÃO DAS BIBLIOTECAS
# -----------------------------------------------------------------------------------------------#

import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook

# -----------------------------------------------------------------------------------------------#
                                       # CRIAÇÃO DAS FUNÇÕES
# -----------------------------------------------------------------------------------------------#

# Função para criar o arquivo Excel se não existir
def criar_arquivo_excel():
    try:
        load_workbook('CadastroArrecadacoes.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Controle"
        ws.append(["Código", "Nome", "Tipo de Arrecadação", "Dia", "Mês", "Ano", "Valor"])
        wb.save('CadastroArrecadacoes.xlsx')

# Função para carregar os dados do Excel para a TreeView
def carregar_dados():
    try:
        wb = load_workbook('CadastroArrecadacoes.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            treeview.insert('', 'end', values=row)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar dados: {e}")

# Função para gravar os dados no Excel
def gravar_dados():
    if (codigo_var.get() == "" or nome_var.get() == "" or tipo_var.get() == "" or
        dia_var.get() == "" or mes_var.get() == "" or ano_var.get() == "" or
        valor_var.get() == ""):
        messagebox.showwarning("Aviso", "Todos os campos devem ser preenchidos.")
        return

    try:
        wb = load_workbook('CadastroArrecadacoes.xlsx')
        ws = wb.active
        ws.append([
            codigo_var.get(), nome_var.get(), tipo_var.get(),
            dia_var.get(), mes_var.get(), ano_var.get(),
            valor_var.get()
        ])
        wb.save('CadastroArrecadacoes.xlsx')
        atualizar_treeview()
        limpar_campos()
    except Exception as e:
        messagebox.showwarning("Aviso", "Planilha Cadastro Arrecadacoes Aberta! Os dados não podem ser Gravados! Favor fechar a Planilha.")
        #messagebox.showerror("Erro", f"Erro ao gravar dados: {e}") 

# Função para atualizar os dados no Excel
def atualizar_dados():
    if not treeview.selection():
        messagebox.showwarning("Aviso", "Selecione o item! Efetue a Alteração no Campo desejado! Clique no Botão Atualizar.")
        return

    selected_item = treeview.selection()[0]
    values = treeview.item(selected_item, 'values')

    if (codigo_var.get() == "" or nome_var.get() == "" or tipo_var.get() == "" or
        dia_var.get() == "" or mes_var.get() == "" or ano_var.get() == "" or
        valor_var.get() == ""):
        messagebox.showwarning("Aviso", "Todos os campos devem ser preenchidos.")
        return

    try:
        wb = load_workbook('CadastroArrecadacoes.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == values[0]:  # Código único
                row[0].value = codigo_var.get()
                row[1].value = nome_var.get()
                row[2].value = tipo_var.get()
                row[3].value = dia_var.get()
                row[4].value = mes_var.get()
                row[5].value = ano_var.get()
                row[6].value = valor_var.get()
                break
        wb.save('CadastroArrecadacoes.xlsx')
        atualizar_treeview()
        limpar_campos()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao atualizar dados: {e}")

# Função para excluir os dados no Excel
def excluir_dados():
    if not treeview.selection():
        messagebox.showwarning("Aviso", "Selecione o item! Clique no Botão Excluir.")
        return

    selected_item = treeview.selection()[0]
    values = treeview.item(selected_item, 'values')

    try:
        wb = load_workbook('CadastroArrecadacoes.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == values[0]:  # Código único
                ws.delete_rows(row[0].row)
                break
        wb.save('CadastroArrecadacoes.xlsx')
        atualizar_treeview()
        limpar_campos()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao excluir dados: {e}")

# Função para atualizar a TreeView com os dados do Excel
def atualizar_treeview():
    for item in treeview.get_children():
        treeview.delete(item)
    carregar_dados()

# Função para limpar os campos de entrada
def limpar_campos():
    codigo_var.set("")
    nome_var.set("")
    tipo_var.set("")
    dia_var.set("")
    mes_var.set("")
    ano_var.set("")
    valor_var.set("")

# Função para preencher os campos de entrada com os dados selecionados
def selecionar_item(event):
    selected_item = treeview.selection()[0]
    values = treeview.item(selected_item, 'values')
    codigo_var.set(values[0])
    nome_var.set(values[1])
    tipo_var.set(values[2])
    dia_var.set(values[3])
    mes_var.set(values[4])
    ano_var.set(values[5])
    valor_var.set(values[6])
    
# -----------------------------------------------------------------------------------------------#
                                      # CRIAÇÃO INTERFACE GRÁFICA
# -----------------------------------------------------------------------------------------------#    

# Configura a interface gráfica
root = tk.Tk()
root.title("Controle Financeiro de Arrecadações Asapac São João del Rei")
root.geometry("885x610+200+50") # Aumanta o tamanho e posiciona a janela na tela

# -----------------------------------------------------------------------------------------------#
                                    # CRIAÇÃO TÍTULO DA APLICAÇÃO
# -----------------------------------------------------------------------------------------------#

# Título
titulo = tk.Label(root, text="Controle Financeiro de Arrecadações Asapac São João del Rei",  fg="blue", font=('Helvetica', 18, 'bold'))
titulo.pack(pady=10)

# -----------------------------------------------------------------------------------------------#
                                   # MANIPULAÇÃO DE DADOS NO EXCEL
# -----------------------------------------------------------------------------------------------#

# Cria e carregar o arquivo Excel
criar_arquivo_excel()

# -----------------------------------------------------------------------------------------------#
                                      # CRIAÇÃO FRAME DE ENTRADA
# -----------------------------------------------------------------------------------------------#

# Frame para os campos de entrada
frame_entrada = tk.Frame(root)
frame_entrada.pack(pady=10)


# -----------------------------------------------------------------------------------------------#
                               # CRIAÇÃO DOS CAMPOS PARA ENTRADA DE DADOS
# -----------------------------------------------------------------------------------------------#

# Labels e Entradas de Dados
labels = ["Código", "Nome", "Tipo de Arrecadação", "Dia", "Mês", "Ano", "Valor"]
entrada_vars = [tk.StringVar() for _ in labels]

for i, (label, var) in enumerate(zip(labels, entrada_vars)):
    tk.Label(frame_entrada, text=f"{label}:", font=("Helvetica", 10),).grid(row=i, column=0, padx=5, pady=5, sticky='w')
    tk.Entry(frame_entrada, textvariable=var, width=50,).grid(row=i, column=1, padx=5, pady=5, sticky='w')

# Variáveis de entrada para os campos
codigo_var, nome_var, tipo_var, dia_var, mes_var, ano_var, valor_var = entrada_vars

# -----------------------------------------------------------------------------------------------#
                                      # CRIAÇÃO TREEVIEW
# -----------------------------------------------------------------------------------------------#

# TreeView para mostrar os dados
frame_treeview = tk.Frame(root)
frame_treeview.pack(pady=10)

colunas = ("Código", "Nome", "Tipo de Arrecadação", "Dia", "Mês", "Ano", "Valor")
treeview = ttk.Treeview(frame_treeview, columns=colunas, show='headings')
treeview.pack(side='left', fill='both', expand=True)

# Configura a largura e centralização das colunas
for coluna in colunas:
    treeview.heading(coluna, text=coluna, anchor='center')
    treeview.column(coluna, width=120, anchor='center')  # Ajuste o valor de width conforme necessário

scroll_y = ttk.Scrollbar(frame_treeview, orient='vertical', command=treeview.yview)
scroll_y.pack(side='right', fill='y')
treeview.configure(yscrollcommand=scroll_y.set)

treeview.bind("<ButtonRelease-1>", selecionar_item)

# -----------------------------------------------------------------------------------------------#
                                      # CRIAÇÃO DOS BOTÕES
# -----------------------------------------------------------------------------------------------#

# Botões
frame_botoes = tk.Frame(root)
frame_botoes.pack(pady=5) # Distancia entre os Botões e a View

# Configurações Formatação Botões
button_config = {
    "padx": 10,     # Espaço horizontal dentro do Botão
    "pady": 5,      # Espaço vertical dentro do Botão
    "width": 20,    # Largura do Botão
    "height": 0,    # Altura do Botão
    "font": ('Helvetica', 10)  # Tamanho da fonte
}

tk.Button(frame_botoes, text="Gravar", command=gravar_dados, **button_config).grid(row=0, column=0, padx=10, pady=5)
tk.Button(frame_botoes, text="Atualizar", command=atualizar_dados, **button_config).grid(row=0, column=1, padx=10, pady=5)
tk.Button(frame_botoes, text="Limpar", command=limpar_campos, **button_config).grid(row=0, column=2, padx=10, pady=5)
tk.Button(frame_botoes, text="Excluir", command=excluir_dados, **button_config).grid(row=0, column=3, padx=10, pady=5)

# Carregar os dados na TreeView ao iniciar
atualizar_treeview()

root.mainloop()
